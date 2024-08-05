import sqlite3
import openpyxl
from bs4 import BeautifulSoup
import chardet

def ENCOURS(param, curseur):
    workbook = openpyxl.load_workbook(str(param))
    feuille = workbook.active

    for row in feuille.iter_rows(min_row=8, values_only=True): 
        if row[2] != None:
            curseur.execute('''
                            INSERT INTO encours (
                                caisse, code_client, rib, nom, adresse, produit, date, duree, montant, 
                                encours, dav, dsp, capital, interet, penalite, retard, agent
                            ) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                            (
                                row[1], row[2], row[3], row[4], row[5], row[9], row[13], row[16], row[19], 
                                row[21], row[24], row[25], row[40], row[41], row[42], row[44], row[50]
                            )
                        )

def RADIE(param, curseur):
    workbook = openpyxl.load_workbook(str(param))
    feuille = workbook.active

    for row in feuille.iter_rows(min_row=8, values_only=True): 
        if row[2] != None:
            curseur.execute('''
                            INSERT INTO radie (
                                caisse, code_client, rib, nom, adresse, produit, date, duree, montant, 
                                capital_perte, date_perte, retard, capital, interet, penalite, reprise, agent
                            ) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                            (
                                row[1], row[2], row[3], row[4], row[5], row[9], row[13], row[16], row[19], 
                                row[21], row[22], row[23], row[24], row[25], row[26], row[31], row[32]
                            )
                        )
            
def DEPOT(param, curseur):
    # Lire le fichier en mode binaire
    with open(param, 'rb') as f:
        raw_data = f.read()

    # Détecter l'encodage
    result = chardet.detect(raw_data)
    encoding = result['encoding']

    with open(param, 'r', encoding=encoding) as file:
        html_content = file.read()
    
    soup = BeautifulSoup(html_content, 'html.parser')

    rows = soup.find_all('tr')

    data = []

    for row in rows:
        cells = row.find_all('td', class_=['EVEN', 'ODD'])
        
        if len(cells) > 0:
            col1 = cells[0].get_text(strip=True) 
            col2 = cells[3].get_text(strip=True) 
            col3 = cells[6].get_text(strip=True) 
            col4 = cells[7].get_text(strip=True)
            col5 = cells[8].get_text(strip=True)
            col6 = cells[10].get_text(strip=True)
            data.append((col1, col2, col3, col4[:-8], col5, col6))
   
    curseur.executemany('''
    INSERT INTO depot (cancel, op, caisse, client, rib, montant)
    VALUES (?, ?, ?, ?, ?, ?)
    ''', data)


    
def importExcelSql(param, test):
    connexion = sqlite3.connect('base.db')
    curseur = connexion.cursor()

    # Vider la table
    curseur.execute(f"DELETE FROM {test}")

    # Réinitialiser la séquence auto-incrémentée
    curseur.execute(f"DELETE FROM sqlite_sequence WHERE name='{test}'")

    connexion.commit()

    try:
        if test == 'encours':
            ENCOURS(param, curseur)
        elif test == 'depot':
            DEPOT(param, curseur)
        else:
            RADIE(param, curseur)   

    except Exception as e:
        print(f"Erreur: {e}")
        connexion.rollback()
    

    connexion.commit()
    connexion.close()

class BDD_SQL:
    def __init__(self):
        self.conn = sqlite3.connect('base.db')
        self.cursor = self.conn.cursor()

    def startBdd(self):
        self.cursor = self.conn.cursor()

    def bddEncours(self):
        query = "SELECT * FROM encours"
        self.cursor.execute(query)

        result = self.cursor.fetchall()

        return result
    
    def bddRadie(self):
        query = "SELECT * FROM radie"
        self.cursor.execute(query)

        result = self.cursor.fetchall()

        return result
    
    def getData(self, bdd, ref=''):
        query = f"SELECT * FROM {bdd}" if ref == '' else f"SELECT * FROM {bdd} WHERE {ref}"
        self.cursor.execute(query)

        result = self.cursor.fetchall()

        return result
    
    def closeBdd(self):

        # Fermer la connexion
        self.cursor.close()
        self.conn.close()

# db = BDD_SQL()
# print(db.bddEncours())
# db.closeBdd()